from openpyxl import load_workbook
from datetime import datetime, timedelta
import sqlite3

def group_index(group):
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    index = 0
    for i in range(1, 500):
        if ws.cell(row=2, column=i).value == str(group):
            index = i
            break
    return index

def is_even_current():
    first_week = datetime(2023, 2, 5)
    todaydate = datetime.today()
    amountDays = todaydate - first_week
    if (amountDays.days // 7) % 2 == 0:
        return False
    else:
        return True

def current_week_timetable(group):
    index = group_index(str(group))
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    str_output = ""
    start_index = 0
    if is_even_current():
        start_index = 5
    else:
        start_index = 4
    for j in range(start_index, 88, 2):
        if ws.cell(row=j, column=1).value != None:
            str_output += str(ws.cell(row=j, column=1).value) + "\n"
        if ws.cell(row=j, column=index).value != "":
            # Время начало пары
            tmp = j
            if ws.cell(row=j, column=3).value == None:
                tmp = j - 1
            str_output += str(ws.cell(row=tmp, column=3).value) + " || "
            str_output += str(ws.cell(row=tmp, column=4).value) + " || "
            # Вывод названия предмета
            str_output += str(ws.cell(row=j, column=index).value) + " || "
            # Лекция или практика
            str_output += str(ws.cell(row=j, column=index + 1).value) + " || "
            # Аудитория
            if ws.cell(row=j, column=index + 2).value != "":
                str_output += str(ws.cell(row=j, column=index + 2).value) + "\n"
            else:
                str_output += "\n"
    return str_output

