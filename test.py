from openpyxl import load_workbook
from datetime import datetime, timedelta

# Функция перевода даты из строки в datetime

def from_string_datetime(str_date):
    arr = []
    if "." in str_date:
        array_of_dates = str_date.split(".")
        day = int(array_of_dates[0])
        month = int(array_of_dates[1])
        year = int(array_of_dates[2])
        return datetime(year, month, day)
    elif "/" in str_date:
        array_of_dates = str_date.split("/")
        day = int(array_of_dates[0])
        month = int(array_of_dates[1])
        year = int(array_of_dates[2])
        return datetime(year, month, day)
    elif " " in str_date:
        array_of_dates = str_date.split()
        day = int(array_of_dates[0])
        month = int(array_of_dates[1])
        year = int(array_of_dates[2])
        return datetime(year, month, day).date
    else:
        return "error"

# Вспомогательная функция вывода списка состоящего из названий групп

def group_arr():
    array_of_groups = []
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    for i in range(1, 500):
        if str(ws.cell(row=2, column=i).value)[0] == "Б":
            array_of_groups.append(ws.cell(row=2, column=i).value)
    return array_of_groups


# Функция проверки наличия группы

def is_group_aviable(str_group):
    array_of_groups = group_arr()
    flag = False
    for each in array_of_groups:
        if str(str_group) == str(each):
            flag = True
            break
    return flag

# Вспомогательная функция определяющая четность или нечетность актуальной недели
def is_even(str_date):
    first_week = datetime(2023, 2, 5)
    todaydate = from_string_datetime(str_date)
    amountDays = todaydate - first_week
    if (amountDays.days // 7) % 2 == 0:
        return False
    else:
        return True

def is_even_current():
    first_week = datetime(2023, 2, 5)
    todaydate = datetime.today()
    amountDays = todaydate - first_week
    if (amountDays.days // 7) % 2 == 0:
        return False
    else:
        return True

# Вспомогательная функция определения индекса группы

def group_index(group):
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    index = 0
    for i in range(1, 500):
        if ws.cell(row=2, column=i).value == str(group):
            index = i
            break
    return index

# Функция определения расписания на актуальную неделю

def current_week_timetable(group):
    index = group_index(str(group))
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    str_output = ""
    for j in range(4, 88):
        if ws.cell(row=j, column=1).value != None:
            str_output += str(ws.cell(row=j, column=1).value) + "\n"
        if ws.cell(row=j, column=index).value != "":
            # Четность нечетность
            if is_even_current():
                if ws.cell(row=j, column=5).value == "II":
                    # Время начало пары
                    tmp = j
                    if ws.cell(row=j, column=3).value == None:
                        tmp = j - 1
                    str_output += str(ws.cell(row=tmp, column=3).value) + " || "
                    str_output += str(ws.cell(row=tmp, column=4).value) + " || "
                    # Вывод названия предмета
                    str_output += str(ws.cell(row=j, column=index).value) + " || "
                    # Лекция или практика
                    str_output += str(ws.cell(row=j, column=index + 1).value) + " || "
                    # Аудитория
                    if ws.cell(row=j, column=index + 2).value != "":
                        str_output += str(ws.cell(row=j, column=index + 2).value) + "\n"
                    else:
                        str_output += "\n"
            else:
                if ws.cell(row=j, column=5).value == "I":
                    # Время начало пары
                    tmp = j
                    if ws.cell(row=j, column=3).value == None:
                        tmp = j - 1
                    str_output += str(ws.cell(row=tmp, column=3).value) + " || "
                    str_output += str(ws.cell(row=tmp, column=4).value) + " || "
                    # Вывод названия предмета
                    str_output += str(ws.cell(row=j, column=index).value) + " || "
                    # Лекция или практика
                    str_output += str(ws.cell(row=j, column=index + 1).value) + " || "
                    # Аудитория
                    if ws.cell(row=j, column=index + 2).value != "":
                        str_output += str(ws.cell(row=j, column=index + 2).value) + "\n"
                    else:
                        str_output += "\n"
    return str_output

# Вспомогательная функция определения дня недели по его номеру

def current_day_of_the_week(str_date):
    datetime_date = from_string_datetime(str_date)
    week_num = datetime_date.isoweekday()
    if week_num == 1:
        return "ПОНЕДЕЛЬНИК"
    elif week_num == 2:
        return "ВТОРНИК"
    elif week_num == 3:
        return "СРЕДА"
    elif week_num == 4:
        return "ЧЕТВЕРГ"
    elif week_num == 5:
        return "ПЯТНИЦА"
    elif week_num == 6:
        return "СУББОТА"

# Функция определения расписания на конкретный день

def current_day_timetable(group, str_date):
    index = group_index(str(group))
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    str_output = ""
    flag=False
    for j in range(4, 88):
        if ws.cell(row=j, column=1).value != None:
            flag = False
        if ws.cell(row=j, column=1).value == current_day_of_the_week(str_date):
            flag = True
        if flag:
            if ws.cell(row=j, column=1).value != None:
                str_output += str(ws.cell(row=j, column=1).value) + "\n"
            if ws.cell(row=j, column=index).value != "":
                # Четность нечетность
                if is_even(str_date) and ws.cell(row=j, column=5).value == "II":
                        # Время начало пары
                        tmp = j
                        if ws.cell(row=j, column=3).value == None:
                            tmp = j - 1
                        str_output += str(ws.cell(row=tmp, column=3).value) + " || "
                        str_output += str(ws.cell(row=tmp, column=4).value) + " || "
                        # Вывод названия предмета
                        str_output += str(ws.cell(row=j, column=index).value) + " || "
                        # Лекция или практика
                        str_output += str(ws.cell(row=j, column=index + 1).value) + " || "
                        # Аудитория
                        if ws.cell(row=j, column=index + 2).value != "":
                            str_output += str(ws.cell(row=j, column=index + 2).value) + "\n"
                        else:
                            str_output += "\n"
                elif not(is_even(str_date)) and ws.cell(row=j, column=5).value == "I":
                        # Время начало пары
                        tmp = j
                        if ws.cell(row=j, column=3).value == None:
                            tmp = j - 1
                        str_output += str(ws.cell(row=tmp, column=3).value) + " || "
                        str_output += str(ws.cell(row=tmp, column=4).value) + " || "
                        # Вывод названия предмета
                        str_output += str(ws.cell(row=j, column=index).value) + " || "
                        # Лекция или практика
                        str_output += str(ws.cell(row=j, column=index + 1).value) + " || "
                        # Аудитория
                        if ws.cell(row=j, column=index + 2).value != "":
                            str_output += str(ws.cell(row=j, column=index + 2).value) + "\n"
                        else:
                            str_output += "\n"
    return str_output

print(current_day_timetable("БСБО-10-21", "11.04.2023"))
print(current_week_timetable("БСБО-10-21"))




