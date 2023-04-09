from openpyxl import load_workbook

workbook = load_workbook('temp.xlsx')
first_sheet = workbook.get_sheet_names()[0]
ws = workbook.get_sheet_by_name(first_sheet)

index = 0
for i in range(1, 500):
    if ws.cell(row=2, column=i).value == "БСБО-10-21":
        index = i
        break
print(index)
k_num = 0
if ws.cell(row=index-1, column=4).value == None:
    for j in range(4, 88):
        flag = False
        for i in range(index - 1, index + 4):
            if i == index - 1:
                flag = True
                k = 0
                k_num += 1
                if k_num % 2 != 0:
                    print("I", end=" ")
                else:
                    print("II", end=" ")
            if flag and k < 4:
                k += 1
                print(ws.cell(row=j, column=i).value, end=" ")
            elif flag and k == 4:
                k += 1
                print(ws.cell(row=j, column=i).value, end="\n")
            else:
                k = 0
                flag = False
                continue
else:
    for j in range(4, 88):
        flag = False
        for i in range(index-5, index+4):
            if ws.cell(row=j, column=i).value == "I" or ws.cell(row=j, column=i).value == "II":
                flag = True
                k = 0
            if flag and k < 4:
                k += 1
                print(ws.cell(row=j, column=i).value, end=" ")
            elif flag and k == 4:
                k += 1
                print(ws.cell(row=j, column=i).value, end="\n")
            else:
                k = 0
                flag = False
                continue
