import sqlite3
from openpyxl import load_workbook
import re

def group_index(group):
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    index = 0
    for i in range(1, 500):
        if ws.cell(row=2, column=i).value == str(group):
            index = i
            break
    return index

def group_arr():
    array_of_groups = []
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    sampleGroup = '\w{4}-\d\d-\d\d'
    for i in range(1, 500):
        match = re.match(sampleGroup, str(ws.cell(row=2, column=i).value))
        if match != None:
            array_of_groups.append(ws.cell(row=2, column=i).value)
    return array_of_groups

def group_to_bd(user_id, group):
    sqlite_connection = sqlite3.connect('Timetable_DB.db')
    cursor = sqlite_connection.cursor()
    question_to_database = cursor.execute(f"SELECT user_tg_id FROM User WHERE user_tg_id = '{user_id}'")
    if question_to_database.fetchall() == []:
        insert_user_id = f"INSERT INTO User (user_tg_id, group_num) VALUES ('{user_id}', '{group}')"
        cursor.execute(insert_user_id)
    else:
        update_user_id = f"UPDATE User SET group_num = '{group}' WHERE user_tg_id = '{user_id}'"
        cursor.execute(update_user_id)
    sqlite_connection.commit()

def parse_group_to_database(group):
    workbook = load_workbook('temp.xlsx')
    ws = workbook[workbook.sheetnames[0]]
    sqlite_connection = sqlite3.connect('Timetable_DB.db')
    cursor = sqlite_connection.cursor()
    index = group_index(str(group))
    weekday_name = ""
    for j in range(4, 88):
        even = ""
        time = ""
        subject_name = ""
        lect_or_prac = ""
        place = ""
        teacher = ""
        if ws.cell(row=j, column=1).value != None:
            weekday_name = str(ws.cell(row=j, column=1).value)
        if ws.cell(row=j, column=index).value != "" and ws.cell(row=j, column=index).value is not None:
            # Четность нечетность
            even = str(ws.cell(row=j, column=5).value)
            # Время начало пары
            tmp = j
            if ws.cell(row=j, column=3).value == None:
                tmp = j - 1
            time = str(ws.cell(row=tmp, column=3).value) + " "
            time += str(ws.cell(row=tmp, column=4).value)
            # Вывод названия предмета
            subject_name = str(ws.cell(row=j, column=index).value)
            # Лекция или практика
            lect_or_prac = str(ws.cell(row=j, column=index + 1).value)
            # Препод
            if ws.cell(row=j, column=index + 2).value != "" and ws.cell(row=j, column=index + 2).value != None:
                teacher = str(ws.cell(row=j, column=index + 2).value)
            else:
                teacher = ""
            if ws.cell(row=j, column=index + 3).value != "" and ws.cell(row=j, column=index + 3).value != None:
                place = str(ws.cell(row=j, column=index + 3).value)
            else:
                place = ""
            question_to_database = f"INSERT INTO timetable (group_num, even, day_of_week, interval_pairs, name, type, place, teacher_name) " \
                                   f"VALUES ('{group}', '{even}', '{weekday_name}', '{time}', '{subject_name}', '{lect_or_prac}', '{place}', '{teacher}')"
            cursor.execute(question_to_database)
            sqlite_connection.commit()